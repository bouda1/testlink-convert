#!/usr/bin/python3

import xml.dom.minidom
from colored import fg, attr
from openpyxl import Workbook
import TestParser
import sys

parser = TestParser.TestParser()
if len(sys.argv) <= 1:
    print("Usage: convert.py xmlfile")
    exit(2)
xml_file = sys.argv[1]

# Open XML document using minidom parser
#  DOMTree = xml.dom.minidom.parse("/home/david/test.xml")
DOMTree = xml.dom.minidom.parse("test.xml")
collection = DOMTree.documentElement
if collection.hasAttribute("testsuite"):
    print("Root element : %s" % collection.getAttribute("testsuite"))

# Get all the movies in the collection
tests = collection.getElementsByTagName("testcase")


# Print detail of each movie.
workbook = Workbook()
sheet = workbook.active

sheet["A1"] = "Test"
sheet["B1"] = "Summary"
sheet["C1"] = "Preconditions"
sheet["D1"] = "Steps"
sheet["E1"] = "Results"

start_row = 1
max_row = 1
row = 1
count = 0
for test in tests:
    start_row = max_row + 1
    print("\n******************************* Test ***************************")
    if test.hasAttribute("name"):
        sheet.cell(row=start_row, column=1).value = test.getAttribute("name")
        print("Test %s" % test.getAttribute("name"))
        if start_row > max_row:
            max_row = start_row

    if test.getElementsByTagName("summary")[0].firstChild is not None:
        summary = test.getElementsByTagName("summary")[0].firstChild.data
        parser.clear_list()
        parser.feed(summary)
        lst = parser.get_list()
        parser.clear_list()
        # lst = extract_test(summary)
        row = start_row
        for t in lst:
            sheet.cell(row=row, column=2).value = t
            row += 1
        print(fg('green') + "Summary: " + "\n".join(lst) + attr('reset'))
        if row > max_row:
            max_row = row

    if test.getElementsByTagName("preconditions")[0].firstChild is not None:
        summary = test.getElementsByTagName("preconditions")[0].firstChild.data
        parser.clear_list()
        parser.feed(summary)
        lst = parser.get_list()
        # lst = extract_test(summary)
        row = start_row
        for t in lst:
            sheet.cell(row=row, column=3).value = t
            row += 1
        print(fg('red') + "Preconditions: " + "\n".join(lst) + attr('reset'))
        if row > max_row:
            max_row = row

    steps = test.getElementsByTagName("steps")
    if len(steps) > 0:
        steps = steps[0].getElementsByTagName("step")

        row = start_row
        print()
        for s in steps:
            step_number = int(s.getElementsByTagName("step_number")[0].firstChild.data)
            # lst = extract_test(summary)
            print(fg("yellow"), end="")
            actions = s.getElementsByTagName("actions")[0].firstChild
            if actions is not None:
                # actions = extract_test(s.getElementsByTagName("actions")[0].firstChild.data)
                parser.clear_list()
                parser.feed(actions.data)
                actions = parser.get_list()
                print(step_number, "-", "\n".join(actions), end="")
                sheet.cell(row=row, column=4).value = str(step_number) + " - " + "\n".join(actions)
                if row > max_row:
                    max_row = row
            print(fg("cyan"), end="")
            expected_results = s.getElementsByTagName("expectedresults")[0].firstChild
            if expected_results is not None:
                parser.clear_list()
                parser.feed(expected_results.data)
                expected_results = parser.get_list()
                print("  =>", "\n".join(expected_results), end="")
                sheet.cell(row=row, column=5).value = "\n".join(expected_results)
                if row > max_row:
                    max_row = row
            print(attr('reset'))
            row += 1
    count += 1
workbook.save(filename="./test.xlsx")
